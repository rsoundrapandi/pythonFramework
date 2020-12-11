import openpyxl

# book = openpyxl.load_workbook("\\.xlsx")
# sheet = book.active
# Dict = {}
# cell = sheet.cell(row=1, column=2)
# # print(cell.value)
# sheet.cell(row=2, column=2).value = "Rahul"
#
# # print(sheet.cell(row=2, column=2).value)
# #
# # print(sheet.max_row)
# #
# # print(sheet.max_column)
# #
# # print(sheet['A5'].value)
#
# for i in range(1, sheet.max_row + 1):  # to get rows
#     # if sheet.cell(row=i, column=1).value == "Soundar2":
#
#     for j in range(1, sheet.max_column + 1):  # to get columns
#         # Dict["lastname"]="shetty
#         # print(sheet.cell(row=i, column=j).value)
#
#         Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
#
# print(Dict)


wb = openpyxl.load_workbook('\\.xlsx')
sheet = wb['Sheet1']
final_list = []
for row in range(2, sheet.max_row + 1):
    FirstName = sheet['A' + str(row)].value
    LastName = sheet['B' + str(row)].value

    final_list.append({'FirstName': FirstName, 'LastName': LastName})

print(final_list)


# @pytest.fixture(params=HomePageData.foo())
# def getData(self, request):
#     return request.param
